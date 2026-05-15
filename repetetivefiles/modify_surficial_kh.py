from openpyxl import load_workbook
from copy import copy

# ============================================================
# INPUTS
# ============================================================
infile = r"/mnt/d/Users/abolmaal/modelling/Modflow/Prep/GreatLakes/model_Layers/HK/mDay_Calibrated_surficial_Kh.xlsx"
outfile = r"/mnt/d/Users/abolmaal/modelling/Modflow/Prep/GreatLakes/model_Layers/HK/mDay_Calibrated_surficial_Kh_modified.xlsx"
sheet_name = "GLB_surf_dissolve_merged"

# ============================================================
# ------------------------------------------------------------
# SUGGESTED HYDROFACIES-BASED K VALUES (m/day)
# tuple = (Upper_mday, Middle_mday)
# ------------------------------------------------------------
K_MD = {
    "coarse_glaciofluvial": (25.0, 10.0), # Used for outwash, ice-contact, sand and gravel
    "coarse_proglacial":    (10.0, 4.0),  # I chose this as a conservative coarse value. It is well below the extreme upper values for gravel and clean sand, but still clearly higher than till or fine deposits.
    "dune_sand":            (12.0, 4.0), # Same idea, but a little lower than the cleanest glaciofluvial materials.
    "alluvial":             (6.0, 2.0), # Based on clean sand behavior, but kept moderate. Alluvium can vary a lot, but at basin scale I treated it as mixed sand/silt, not pure gravel.
    "colluvial":            (1.0, 0.3), # 
    "coastal_medium":       (4.0, 1.5), #These were placed lower because coastal and nearshore deposits are often more mixed and less uniformly coarse.
    "nearshore_mixed":      (1.0, 0.3),

    "blanket_moraine":      (0.05, 0.02), # These were chosen to stay consistent with the glacial till, silty sand, and silt/loess ranges.
    "till_sandy":           (0.05, 0.02),  # That part is strongly tied to the glacial till literature
    "till_silty":           (0.01, 0.003),
    "till_clayey":          (0.002, 0.0005),
    "proglacial_fine":      (0.02, 0.005),

    "offshore_fine":        (5.0e-5, 1.0e-5), # this is in a range of unweathered marine clays, which can be low because they act more like confining material 

    "organic":              (0.2, 0.05), # 
    "residual_carbonate":   (0.05, 0.015), # that weathered residual carbonate soils may inherit more permeability than shale-rich residuals, but much less than open coarse sand.
    "residual_sedimentary": (0.02, 0.005), # should be less than open coarse sand 
    "residual_igmet":       (0.002, 0.0005),
    "residual_generic":     (0.01, 0.003),

    "bedrock":              (0.01, 0.002), 
    "volcanic":             (0.5, 0.1),
    "water":                (0.0, 0.0),
    "other":                (0.2, 0.05),
}

QUALIFIER_MULT = {
    "thin": 0.7,
    "discontinuous": 0.7,
    "veneer": 0.7,
    "thick": 1.3,
    "blanket": 1.0,
}

def md_to_ms(k_md):
    return k_md / 86400.0

def classify_hydrofacies(desc):
    if desc is None:
        return None

    d = str(desc).lower()

    if "water" in d:
        return "water"
    if "volcanic" in d:
        return "volcanic"
    if "organic" in d or "peat" in d or "muck" in d:
        return "organic"

    # bedrock / residual
    if "bedrock" in d:
        return "bedrock"
    if "residual" in d and ("dolomite" in d or "limestone" in d or "carbonate" in d):
        return "residual_carbonate"
    if "residual" in d and ("sandstone" in d or "shale" in d or "sedimentary" in d):
        return "residual_sedimentary"
    if "residual" in d and ("igneous" in d or "metamorphic" in d or "granite" in d or "gneiss" in d):
        return "residual_igmet"
    if "residual" in d:
        return "residual_generic"

    # tills
    if "glacial till" in d and "sandy" in d:
        return "till_sandy"
    if "glacial till" in d and "silty" in d:
        return "till_silty"
    if "glacial till" in d and "clayey" in d:
        return "till_clayey"
    if "till" in d and "sandy" in d:
        return "till_sandy"
    if "till" in d and "silty" in d:
        return "till_silty"
    if "till" in d and "clayey" in d:
        return "till_clayey"
    if "till" in d:
        return "till_silty"

    # coarse
    if "glaciofluvial" in d or "ice-contact" in d or "outwash" in d:
        return "coarse_glaciofluvial"
    if "proglacial" in d and "coarse" in d:
        return "coarse_proglacial"

    # fine
    if "offshore" in d or "marine clay" in d or "glaciomarine" in d or "glaciolacustrine" in d:
        return "offshore_fine"
    if "proglacial" in d and "fine" in d:
        return "proglacial_fine"

    # sands / alluvial
    if "eolian" in d or "aeolian" in d or "dune sand" in d:
        return "dune_sand"
    if "alluvial" in d:
        return "alluvial"
    if "colluvial" in d:
        return "colluvial"
    if "coastal zone" in d and "medium" in d:
        return "coastal_medium"
    if "littoral" in d or "nearshore" in d:
        return "nearshore_mixed"

    # mixed glacial
    if "blanket" in d or "moraine" in d or "veneer" in d:
        return "blanket_moraine"

    return "other"

def apply_qualifier_multiplier(desc, upper_md, middle_md):
    d = str(desc).lower()
    factor = 1.0
    for key, mult in QUALIFIER_MULT.items():
        if key in d:
            factor = mult
            break
    return upper_md * factor, middle_md * factor, factor

# ============================================================
# LOAD WORKBOOK
# ============================================================
wb = load_workbook(infile)
ws = wb[sheet_name]

# ============================================================
# ADD NEW COLUMNS ON THE SAME SHEET
# originals stay in D:G
# E = original Upper m/d
# G = original Middle m/d
# ============================================================
start_col = ws.max_column + 1

headers = [
    "Hydrofacies",
    "Suggested Kh (m/d) Upper",
    "Suggested Kh (m/s) Upper",
    "Suggested Kh (m/d) Middle",
    "Suggested Kh (m/s) Middle",
    "Qualifier factor",
    "Note",
]

for i, h in enumerate(headers, start=start_col):
    ws.cell(1, i).value = h

    # copy style from an existing header cell if available
    ref = ws.cell(1, 3)
    ws.cell(1, i).font = copy(ref.font)
    ws.cell(1, i).fill = copy(ref.fill)
    ws.cell(1, i).border = copy(ref.border)
    ws.cell(1, i).alignment = copy(ref.alignment)

# set widths
for i in range(start_col, start_col + len(headers)):
    col_letter = ws.cell(1, i).column_letter
    ws.column_dimensions[col_letter].width = 22

# ============================================================
# PROCESS ROWS
# ============================================================
for r in range(3, ws.max_row + 1):
    desc = ws.cell(r, 3).value

    if desc is None:
        continue
    if isinstance(desc, str) and desc.startswith("* Corresponds to"):
        continue

    hydro = classify_hydrofacies(desc)

    sug_upper_md, sug_middle_md = K_MD[hydro]
    sug_upper_md, sug_middle_md, factor = apply_qualifier_multiplier(
        desc, sug_upper_md, sug_middle_md
    )

    # keep physical consistency
    sug_middle_md = min(sug_middle_md, sug_upper_md)

    sug_upper_ms = md_to_ms(sug_upper_md)
    sug_middle_ms = md_to_ms(sug_middle_md)

    # write suggested values only
    ws.cell(r, start_col + 0).value = hydro
    ws.cell(r, start_col + 1).value = sug_upper_md
    ws.cell(r, start_col + 2).value = sug_upper_ms
    ws.cell(r, start_col + 3).value = sug_middle_md
    ws.cell(r, start_col + 4).value = sug_middle_ms
    ws.cell(r, start_col + 5).value = factor
    ws.cell(r, start_col + 6).value = "Original table preserved on same sheet"

# ============================================================
# FORMATTING
# ============================================================
for r in range(3, ws.max_row + 1):
    ws.cell(r, start_col + 1).number_format = "0.000000"
    ws.cell(r, start_col + 2).number_format = "0.00000000"
    ws.cell(r, start_col + 3).number_format = "0.000000"
    ws.cell(r, start_col + 4).number_format = "0.00000000"
    ws.cell(r, start_col + 5).number_format = "0.00"

wb.save(outfile)
print(f"Saved: {outfile}")